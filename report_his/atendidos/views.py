""" Atendidos views """
from django.db import connection
from django.db import transaction
# Django
from django.contrib.auth.decorators import login_required
from django.urls import reverse_lazy
from django.shortcuts import render
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
from django.http.response import HttpResponse
# Models
from .models import Disa,Red,Microred,Establecimiento,Tipo_reporte,rpt_metales, Anio, Mes
# Forms
# Excel Export
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, Color
# Utilities
from datetime import datetime

def submenu_diresa(request):
    diresas = Disa.objects.all()
    anio = Anio.objects.all().order_by('-id')
    mes = Mes.objects.all()
    context = {
                'diresas': diresas,
                'anio': anio,
                'mes':mes,
                }
    return render(request, 'atendidos/rpt_diresa.html', context)

def submenu_red(request):
    diresas = Disa.objects.all()
    anio = Anio.objects.all().order_by('-id')
    mes = Mes.objects.all()
    context = {
                'diresas': diresas,
                'anio': anio,
                'mes':mes,
                }
    return render(request, 'atendidos/rpt_red.html', context)

def submenu_microred(request):
    diresas = Disa.objects.all()
    anio = Anio.objects.all().order_by('-id')
    mes = Mes.objects.all()
    context = {
                'diresas': diresas,
                'anio': anio,
                'mes':mes,
                }
    return render(request, 'atendidos/rpt_microred.html', context)

def submenu_establecimiento(request):
    diresas = Disa.objects.all()
    anio = Anio.objects.all().order_by('-id')
    mes = Mes.objects.all()
    context = {
                'diresas': diresas,
                'anio': anio,
                'mes':mes,
                }
    return render(request, 'atendidos/rpt_establecimiento.html', context)

# Create your views here.
@login_required
def diresa(request):
    diresas = Disa.objects.all()
    anio = Anio.objects.all().order_by('-id')
    mes = Mes.objects.all()
    context = {
                'diresas': diresas,
                'anio': anio,
                'mes':mes,
                }
    return render(request, 'atendidos/base.html', context)

@login_required
def redes(request):
    diresa = request.GET.get('diresa_selected')
    redes = Red.objects.filter(disa=diresa)
    context = {'redes': redes, 'is_htmx': True }
    return render(request, 'atendidos/partials/redes.html', context)

@login_required
def microredes(request):
    red = request.GET.get('redes_selected')
    microredes = Microred.objects.filter(cod_enlace_r=red)
    context = {'microredes': microredes }
    return render(request, 'atendidos/partials/microredes.html', context)

@login_required
def establecimientos(request):
    microred = request.GET.get('microredes_selected')
    establecimientos = Establecimiento.objects.filter(cod_enlace_mr=microred)
    context = {'establecimientos': establecimientos}
    return render(request, 'atendidos/partials/establecimientos.html', context)   

@login_required
def rpt_atendidos(request):
    anio = request.GET.get('diresa_selected')
    mes_inicio = request.GET.get('redes_selected')
    mes_fin = request.GET.get('microredes_selected')
    red = request.GET.get('establecimientos_selected')

    cursor = connection.cursor()
    cursor.execute("EXEC sp_rpt_metales1 @Anio = %s, @mes_inicio = %s, @mes_fin = %s, @diresa =%s, @cod_red =%s,@cod_microred=%s,@Id_Establecimiento=%s,@provincia=%s,@distrito=%s", ('2020','5','5','00','','','','',''))
    results = cursor.fetchall()
    print(results)
    # manejar los resultados aquí
    return HttpResponse("Llamada al procedimiento almacenado exitosa", results)
    
    #atendidos = ConsAdolescente.objects.all();
    atendidos = rpt_metales.objects.filter(anio = anio, mes__range=(mes_inicio, mes_fin));
    wb = Workbook()
    ws = wb.active
    ws['B1'] = 'REPORTE DE ATENDIDOS'
    
    ws.merge_cells('B1:E1')
    #Cabecera
    ws['B3'] = 'ID_CITA'
    ws['C3'] = 'CODIGO_ITEM'
    ws['D3'] = 'TIPO DIAGNOSTICO'
    ws['E3'] = 'VALOR LAB'
    
    contador = 4       

    for atendido in atendidos:
        ws.cell(row = contador, column = 2).value = atendido.anio
        ws.cell(row = contador, column = 3).value = atendido.mes
        ws.cell(row = contador, column = 4).value = atendido.id_establecimiento
        ws.cell(row = contador, column = 5).value = atendido.nombre_establecimiento
        contador+=1
    
    nombre_archivo = "Reporte_atendido.xlsx"
    response = HttpResponse(content_type = "application/ms-excel")
    contenido = "attachment; filename = {0}".format(nombre_archivo)
    response["Content-Disposition"]  = contenido
    wb.save(response)
    return response



